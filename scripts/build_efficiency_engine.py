"""
One-time structural migration: replaces the workbook's 'Advanced Efficiency Metrics' tab
(single-season snapshot, Z-scored directly) with a 5-section Multi-Year Efficiency Engine
that mirrors 'YoY Baseline Engine's own pattern exactly -- 3-yr decay-weighted average,
blended with extra last-year emphasis, then regressed toward the league mean -- run once
per efficiency metric instead of once for PPG. See claude_code_spec_efficiency_engine.md.

Also repoints Team Ratings' Efficiency Adjustment column at the new tab layout.

Usage:
    uv run python scripts/build_efficiency_engine.py "C:\\path\\to\\NFL_Prediction_Model.xlsx"

This is a one-off rebuild script, not part of the regular pull/update pipeline -- run it
once, then use nflverse_pull.update_workbook for ongoing refreshes (once that module is
updated to match this tab's new shape).
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from nflverse_pull.efficiency import compute_team_season_efficiency, fetch_pbp  # noqa: E402
from nflverse_pull.pull import TEAM_NAMES  # noqa: E402

YEARS = [2023, 2024, 2025]
SHEET_NAME = "Advanced Efficiency Metrics"

# Canonical 32-team row order, copied from 'YoY Baseline Engine'!A110:A141 -- keeps every
# tab in the workbook walking teams in the same order.
TEAM_ORDER = [
    "Buffalo Bills", "Miami Dolphins", "New England Patriots", "New York Jets",
    "Baltimore Ravens", "Cincinnati Bengals", "Cleveland Browns", "Pittsburgh Steelers",
    "Houston Texans", "Indianapolis Colts", "Jacksonville Jaguars", "Tennessee Titans",
    "Denver Broncos", "Kansas City Chiefs", "Las Vegas Raiders", "Los Angeles Chargers",
    "Dallas Cowboys", "New York Giants", "Philadelphia Eagles", "Washington Commanders",
    "Chicago Bears", "Detroit Lions", "Green Bay Packers", "Minnesota Vikings",
    "Atlanta Falcons", "Carolina Panthers", "New Orleans Saints", "Tampa Bay Buccaneers",
    "Arizona Cardinals", "Los Angeles Rams", "San Francisco 49ers", "Seattle Seahawks",
]
assert set(TEAM_ORDER) == set(TEAM_NAMES.values()), "TEAM_ORDER must cover all 32 teams"

# Six scored metrics, in the exact order the existing Model Assumptions weights (C24:C29)
# already assume. sign_flip=True means "lower is better" (a defense/allowed metric) --
# Z-scored as (league_avg - value)/std instead of (value - league_avg)/std, same convention
# the tab's old Section 3 already used.
METRICS = [
    {"key": "epa_off", "label": "EPA Off", "csv_col": "EPA/Play (Off)",
     "sec1_col": "C", "sign_flip": False, "weight_cell": "$C$24", "fmt": "0.000"},
    {"key": "epa_def", "label": "EPA Def", "csv_col": "EPA/Play Allowed (Def)",
     "sec1_col": "D", "sign_flip": True, "weight_cell": "$C$25", "fmt": "0.000"},
    {"key": "sr_off", "label": "SR Off", "csv_col": "Success Rate (Off)",
     "sec1_col": "E", "sign_flip": False, "weight_cell": "$C$26", "fmt": "0.0%"},
    {"key": "sr_def", "label": "SR Def", "csv_col": "Success Rate Allowed (Def)",
     "sec1_col": "F", "sign_flip": True, "weight_cell": "$C$27", "fmt": "0.0%"},
    {"key": "nya_off", "label": "NY/A Off", "csv_col": "NY/A (Off)",
     "sec1_col": "G", "sign_flip": False, "weight_cell": "$C$28", "fmt": "0.00"},
    {"key": "nya_def", "label": "NY/A Def", "csv_col": "NY/A Allowed (Def)",
     "sec1_col": "H", "sign_flip": True, "weight_cell": "$C$29", "fmt": "0.00"},
]

# --- Styles, copied from the existing tabs (read off, not guessed -- see spec) ----------
TITLE_FONT = Font(name="Arial", size=10, bold=True)
TITLE_FILL = PatternFill("solid", fgColor="FFD9E1F2")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="FF1F4E78")
HEADER_ALIGN = Alignment(wrap_text=True, horizontal="center", vertical="center")
INPUT_FONT = Font(name="Arial", size=10, color="FF0000FF")
FORMULA_FONT = Font(name="Arial", size=10, color="FF000000")
LINK_FONT = Font(name="Arial", size=10, color="FF008000")
NOTE_FONT = Font(name="Arial", size=9, color="FF808080")


def _section_title(ws, row: int, last_col: int, text: str) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL


def _header_row(ws, row: int, headers: list[str], height: float = 39.75) -> None:
    ws.row_dimensions[row].height = height
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN


def build(workbook_path: str) -> None:
    print(f"Pulling {YEARS} play-by-play data...")
    pbp = fetch_pbp(YEARS)
    season_df = compute_team_season_efficiency(pbp)
    by_team_season = {
        (r["Team"], int(r["Season"])): r for r in season_df.to_dict("records")
    }
    missing = [
        (team, yr) for team in TEAM_ORDER for yr in YEARS if (team, yr) not in by_team_season
    ]
    if missing:
        raise ValueError(f"Pulled data is missing rows for: {missing}")

    wb = openpyxl.load_workbook(workbook_path)
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME, index=wb.sheetnames.index("YoY Baseline Engine") + 1)

    ws.column_dimensions["A"].width = 22.0

    # ---- Sheet title -------------------------------------------------------------------
    ws.merge_cells("A1:I1")
    t = ws.cell(row=1, column=1, value=(
        "Advanced Efficiency Metrics (EPA/play, Success Rate, NY/A) -- Multi-Year "
        "Decay-Weighted Engine, mirrors YoY Baseline Engine"
    ))
    t.font = Font(name="Arial", size=12, bold=True)

    # ==== Section 1: Raw 3-Year Data (rows 5-100, cols A-I) ==============================
    _section_title(ws, 3, 9, "Section 1 \u2014 Raw 3-Year Data by Season (from nflverse pbp)")
    _header_row(ws, 4, ["Team", "Season", *[m["label"] for m in METRICS], "PROE"], height=28)

    row = 5
    for team in TEAM_ORDER:
        for yr in YEARS:
            data = by_team_season[(team, yr)]
            ws.cell(row=row, column=1, value=team).font = INPUT_FONT
            c = ws.cell(row=row, column=2, value=yr)
            c.font = INPUT_FONT
            for i, m in enumerate(METRICS):
                cell = ws.cell(row=row, column=3 + i, value=float(data[m["csv_col"]]))
                cell.font = INPUT_FONT
                cell.number_format = m["fmt"]
            proe_cell = ws.cell(row=row, column=9, value=float(data["PROE (Off)"]))
            proe_cell.font = INPUT_FONT
            proe_cell.number_format = "0.0%"
            row += 1
    assert row == 101, f"Section 1 should end at row 101, ended at {row}"

    # ==== Section 2: League Baseline by Season, per metric (rows 102-106, cols A-G) =====
    _section_title(ws, 102, 7, "Section 2 \u2014 League Baseline by Season, per Metric")
    _header_row(ws, 103, ["Season", *[m["label"] for m in METRICS]], height=28)

    for i, yr in enumerate(YEARS):
        r = 104 + i
        c = ws.cell(row=r, column=1, value=yr)
        c.font = INPUT_FONT
        for j, m in enumerate(METRICS):
            col = get_column_letter(3 + j)  # Section 1 metric column, e.g. C for epa_off
            formula = f"=AVERAGEIF($B$5:$B$100,$A{r},{col}$5:{col}$100)"
            cell = ws.cell(row=r, column=2 + j, value=formula)
            cell.font = FORMULA_FONT
            cell.number_format = m["fmt"]

    # ==== Section 3: Per-Team 3-Yr Decay-Weighted, Regressed Baseline (rows 108-141) ====
    # 7 columns per metric (Y-1, Y-2, Y-3, Weighted 3-Yr Avg, Team History, League Baseline,
    # Projected 3-Yr Baseline) x 6 metrics = 42, plus Team (1) and PROE Y-1 passthrough (1)
    # = 44 columns total (A through AR). Exactly replicates YoY Baseline Engine's own
    # Section 3 formula chain, run once per metric instead of once for PPG.
    sec3_last_col = 1 + len(METRICS) * 7 + 1  # 44
    _section_title(
        ws, 108, sec3_last_col,
        "Section 3 \u2014 Per-Team 3-Yr Decay-Weighted, Regressed Baseline, per Metric "
        "(3-yr decay-weighted average is blended with last year's number for extra recency "
        "emphasis, then that combined team-history figure is regressed toward the "
        "prior-year league mean -- same formula chain as YoY Baseline Engine Section 3)",
    )
    headers = ["Team"]
    metric_block_start_col: dict[str, int] = {}  # metric key -> 1-based start column
    col_cursor = 2
    for m in METRICS:
        metric_block_start_col[m["key"]] = col_cursor
        headers += [
            f"{m['label']} Y-1", f"{m['label']} Y-2", f"{m['label']} Y-3",
            f"Weighted\n{m['label']} Avg\n(3-Yr decay)", f"Team History\n{m['label']}",
            f"League Baseline\n{m['label']} (Y-1)", f"Projected 3-Yr\n{m['label']} Baseline",
        ]
        col_cursor += 7
    headers.append("PROE Y-1\n(context only)")
    proe_col = col_cursor  # 44
    _header_row(ws, 109, headers)

    for i, team in enumerate(TEAM_ORDER):
        r = 110 + i
        ws.cell(row=r, column=1, value=team).font = FORMULA_FONT
        for m in METRICS:
            base = metric_block_start_col[m["key"]]
            y1, y2, y3, wavg, th, lb, pb = (get_column_letter(base + k) for k in range(7))
            sec1_col = m["sec1_col"]

            f_y1 = ws.cell(row=r, column=base, value=(
                f"=SUMIFS(${sec1_col}$5:${sec1_col}$100,$A$5:$A$100,$A{r},"
                f"$B$5:$B$100,'Model Assumptions'!$C$18-1)"
            ))
            f_y2 = ws.cell(row=r, column=base + 1, value=(
                f"=SUMIFS(${sec1_col}$5:${sec1_col}$100,$A$5:$A$100,$A{r},"
                f"$B$5:$B$100,'Model Assumptions'!$C$18-2)"
            ))
            f_y3 = ws.cell(row=r, column=base + 2, value=(
                f"=SUMIFS(${sec1_col}$5:${sec1_col}$100,$A$5:$A$100,$A{r},"
                f"$B$5:$B$100,'Model Assumptions'!$C$18-3)"
            ))
            f_wavg = ws.cell(row=r, column=base + 3, value=(
                f"=({y1}{r}*1+{y2}{r}*'Model Assumptions'!$C$20+"
                f"{y3}{r}*('Model Assumptions'!$C$20^2))/"
                f"(1+'Model Assumptions'!$C$20+'Model Assumptions'!$C$20^2)"
            ))
            f_th = ws.cell(row=r, column=base + 4, value=(
                f"={y1}{r}*'Model Assumptions'!$C$22+{wavg}{r}*(1-'Model Assumptions'!$C$22)"
            ))
            sec2_col = get_column_letter(2 + METRICS.index(m))
            f_lb = ws.cell(row=r, column=base + 5, value=(
                f"=INDEX(${sec2_col}$104:${sec2_col}$106,"
                f"MATCH('Model Assumptions'!$C$18-1,$A$104:$A$106,0))"
            ))
            f_pb = ws.cell(row=r, column=base + 6, value=(
                f"={th}{r}*'Model Assumptions'!$C$21+{lb}{r}*(1-'Model Assumptions'!$C$21)"
            ))
            for cell in (f_y1, f_y2, f_y3, f_wavg, f_th, f_lb, f_pb):
                cell.font = FORMULA_FONT
                cell.number_format = m["fmt"]

        proe_formula = ws.cell(row=r, column=proe_col, value=(
            f"=SUMIFS($I$5:$I$100,$A$5:$A$100,$A{r},"
            f"$B$5:$B$100,'Model Assumptions'!$C$18-1)"
        ))
        proe_formula.font = FORMULA_FONT
        proe_formula.number_format = "0.0%"

    # ==== Section 4: League Average & Std. Dev. of Section 3's Projected Baselines ======
    _section_title(ws, 143, 7, "Section 4 \u2014 League Average & Std. Dev. of the 3-Yr Baselines")
    _header_row(ws, 144, ["Stat", *[m["label"] for m in METRICS]], height=20)

    proj_baseline_cols = {
        m["key"]: get_column_letter(metric_block_start_col[m["key"]] + 6) for m in METRICS
    }
    avg_row, std_row = 145, 146
    a = ws.cell(row=avg_row, column=1, value="League Average")
    a.font = FORMULA_FONT
    s = ws.cell(row=std_row, column=1, value="League Std. Dev.")
    s.font = FORMULA_FONT
    for j, m in enumerate(METRICS):
        pcol = proj_baseline_cols[m["key"]]
        avg_cell = ws.cell(row=avg_row, column=2 + j, value=f"=AVERAGE({pcol}$110:{pcol}$141)")
        std_cell = ws.cell(row=std_row, column=2 + j, value=f"=STDEVP({pcol}$110:{pcol}$141)")
        for cell in (avg_cell, std_cell):
            cell.font = FORMULA_FONT
            cell.number_format = m["fmt"]

    # ==== Section 5: Z-Scores and Weighted Efficiency Adjustment (rows 150-181) ==========
    _section_title(
        ws, 148, 9,
        "Section 5 \u2014 Z-Scores and Weighted Efficiency Adjustment (defensive/allowed "
        "metrics are sign-flipped so a higher Z always means better; PROE passed through "
        "for context, not scored)",
    )
    _header_row(
        ws, 149,
        ["Team", *[f"{m['label']}\nZ" + (" (flipped)" if m["sign_flip"] else "")
                    for m in METRICS],
         "Weighted\nEfficiency\nAdjustment (pts)", "PROE\n(context only)"],
    )

    avg_cell_ref = {
        m["key"]: f"${get_column_letter(2 + j)}${avg_row}" for j, m in enumerate(METRICS)
    }
    std_cell_ref = {
        m["key"]: f"${get_column_letter(2 + j)}${std_row}" for j, m in enumerate(METRICS)
    }

    for i, team in enumerate(TEAM_ORDER):
        sec3_row = 110 + i
        r = 150 + i
        ws.cell(row=r, column=1, value=team).font = FORMULA_FONT

        z_cols = []
        for j, m in enumerate(METRICS):
            pcol = proj_baseline_cols[m["key"]]
            avg_ref, std_ref = avg_cell_ref[m["key"]], std_cell_ref[m["key"]]
            if m["sign_flip"]:
                formula = f"=({avg_ref}-{pcol}{sec3_row})/{std_ref}"
            else:
                formula = f"=({pcol}{sec3_row}-{avg_ref})/{std_ref}"
            col = 2 + j
            cell = ws.cell(row=r, column=col, value=formula)
            cell.font = FORMULA_FONT
            cell.number_format = "0.00"
            z_cols.append(get_column_letter(col))

        weighted_terms = " + ".join(
            f"{z_cols[j]}{r}*'Model Assumptions'!{m['weight_cell']}" for j, m in enumerate(METRICS)
        )
        weighted_cell = ws.cell(row=r, column=2 + len(METRICS), value=f"={weighted_terms}")
        weighted_cell.font = FORMULA_FONT
        weighted_cell.number_format = "0.0;(0.0)"

        proe_link = ws.cell(
            row=r, column=3 + len(METRICS), value=f"={get_column_letter(proe_col)}{sec3_row}"
        )
        proe_link.font = FORMULA_FONT
        proe_link.number_format = "0.0%"

    # ---- Closing note --------------------------------------------------------------------
    note_row = 183
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=9)
    note = ws.cell(row=note_row, column=1, value=(
        "This tab replicates the YoY Baseline Engine's decay-weighted, regressed-to-mean "
        "projection methodology, run once per efficiency metric instead of once for PPG. "
        "Section 1 is the raw 3-year 'raw_data' table pulled from nflverse play-by-play "
        "(nflverse_pull.efficiency.compute_team_season_efficiency). Section 2 computes each "
        "metric's league average per season. Section 3 builds each team's 3-yr "
        "decay-weighted average, blends in last-year emphasis, then regresses that combined "
        "team-history figure toward the prior-year league mean -- same formula chain as "
        "'YoY Baseline Engine' Section 3, just run per metric. Section 4 computes the league "
        "average/std-dev of Section 3's Projected 3-Yr Baseline outputs (not the old "
        "single-season raw inputs). Section 5 Z-scores each team's baseline against Section "
        "4 (defensive/allowed metrics sign-flipped so higher Z always means better) and "
        "combines the six scored Z-scores into a single Weighted Efficiency Adjustment using "
        "the same weights as before (Model Assumptions C24:C29) -- no shrink/confidence "
        "discount is applied, since this is now backed by real 3-year history, same as PPG "
        "needs no such discount. PROE (context only) is the team's most recent (Y-1) raw "
        "value, carried through unweighted for the same reason as before: it measures "
        "play-calling tendency, not effectiveness, so it has no 'higher is better' direction "
        "to score."
    ))
    note.font = NOTE_FONT
    note.alignment = Alignment(wrap_text=True, vertical="top")

    # ==== Repoint Team Ratings' Efficiency Adjustment column at the new Section 5 =========
    tr = wb["Team Ratings"]
    weighted_adj_col = get_column_letter(2 + len(METRICS))  # H
    for r in range(3, 3 + len(TEAM_ORDER)):
        cell = tr.cell(row=r, column=13, value=(  # column M
            f"=INDEX('{SHEET_NAME}'!${weighted_adj_col}$150:${weighted_adj_col}$181,"
            f"MATCH(A{r},'{SHEET_NAME}'!$A$150:$A$181,0))"
        ))
        cell.font = LINK_FONT
        cell.number_format = "0.0;(0.0)"

    wb.save(workbook_path)
    print(f"Rebuilt '{SHEET_NAME}' (5 sections, {sec3_last_col} cols in Section 3) and "
          f"repointed Team Ratings column M for {len(TEAM_ORDER)} teams.")
    print(f"Saved to {workbook_path}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print('Usage: uv run python scripts/build_efficiency_engine.py "path/to/workbook.xlsx"')
        sys.exit(1)
    build(sys.argv[1])
