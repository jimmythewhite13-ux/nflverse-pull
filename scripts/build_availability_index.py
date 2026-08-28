"""
Builds Part E of claude_code_spec_qb_index.md: a new "Availability Index" workbook tab
(Injury History + Schedule Density only -- the two of the originally-envisioned 5 injury-
index dimensions that are actually sourceable; ACWR, GPS/load, and wellness surveys are
not, and are deliberately absent here), applied to this season's Starters + Backups.

Requires "QB Index" to already exist (run build_qb_index.py first).

Usage:
    uv run python scripts/build_availability_index.py "C:\\path\\to\\NFL_Prediction_Model.xlsx"
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from nflverse_pull.availability import (  # noqa: E402
    compute_player_injury_history,
    compute_team_game_log,
    fetch_injuries,
    fetch_schedules_with_dates,
)
from nflverse_pull.qb_stats import compute_qb_roles, compute_team_season_qb_stats  # noqa: E402

YEARS = [2023, 2024, 2025]
QB_INDEX_SHEET = "QB Index"
SHEET_NAME = "Availability Index"

TITLE_FONT = Font(name="Arial", size=10, bold=True)
TITLE_FILL = PatternFill("solid", fgColor="FFD9E1F2")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="FF1F4E78")
HEADER_ALIGN = Alignment(wrap_text=True, horizontal="center", vertical="center")
INPUT_FONT = Font(name="Arial", size=10, color="FF0000FF")
FORMULA_FONT = Font(name="Arial", size=10, color="FF000000")
NOTE_FONT = Font(name="Arial", size=9, color="FF808080")
ASSUMPTION_FILL = PatternFill("solid", fgColor="FFFFFF00")


def _section_title(ws, row, last_col, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL


def _header_row(ws, row, headers, height=28):
    ws.row_dimensions[row].height = height
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN


def add_model_assumptions_weights(wb: openpyxl.Workbook) -> None:
    ws = wb["Model Assumptions"]
    ws.merge_cells("A40:D40")
    title = ws.cell(row=40, column=1, value=(
        "Availability Index Weighting (pts per std. dev.; partial index -- Injury History "
        "and Schedule Density only, see 'Availability Index' tab)"
    ))
    title.font = HEADER_FONT
    title.fill = HEADER_FILL

    rows = [
        (41, "Injury History Weight (pts per SD)", 0.5,
         "Fewer recent Out/Doubtful weeks -> higher (better) Z, weighted here. Set to 0 to "
         "exclude this dimension entirely."),
        (42, "Schedule Density Weight (pts per SD)", 0.5,
         "Less recent-trailing-window travel -> higher (better) Z. Note: for a season "
         "opener (Week 1), every team's trailing window is empty by definition (no games "
         "played yet that season) -- this dimension is only meaningful once real games "
         "have been played, e.g. Week 8+."),
    ]
    for row, label, value, note in rows:
        ws.cell(row=row, column=2, value=label)
        c = ws.cell(row=row, column=3, value=value)
        c.font = INPUT_FONT
        c.fill = ASSUMPTION_FILL
        c.number_format = "0.00"
        n = ws.cell(row=row, column=4, value=note)
        n.font = NOTE_FONT
        n.alignment = Alignment(wrap_text=True, vertical="top")


def build(workbook_path: str) -> None:
    print(f"Pulling injuries + schedules for {YEARS}...")
    injuries = fetch_injuries(YEARS)
    sched = fetch_schedules_with_dates(YEARS)

    from nflverse_pull.efficiency import fetch_pbp

    pbp = fetch_pbp(YEARS)
    qb_stats = compute_team_season_qb_stats(pbp)
    roles = compute_qb_roles(qb_stats)
    current_season = max(YEARS)
    current_qbs = (
        roles[(roles["Season"] == current_season) & (roles["Role"].isin(["Starter", "Backup"]))]
        .sort_values(["Team", "Role"])
        .reset_index(drop=True)
    )

    injury_hist = compute_player_injury_history(injuries, set(current_qbs["Player ID"]))
    injury_lookup = injury_hist.set_index("Player ID").to_dict("index")

    game_log = compute_team_game_log(sched)
    print(f"{len(current_qbs)} current Starters/Backups; {len(game_log)} team-game rows pulled.")

    wb = openpyxl.load_workbook(workbook_path)
    add_model_assumptions_weights(wb)

    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME, index=wb.sheetnames.index(QB_INDEX_SHEET) + 1)
    ws.column_dimensions["A"].width = 18.0

    ws.merge_cells("A1:G1")
    t = ws.cell(row=1, column=1, value=(
        "Availability Index -- Injury History + Schedule Density (a PARTIAL index: 2 of "
        "the 5 originally-envisioned dimensions -- ACWR, GPS/load, and wellness-survey data "
        "are not publicly sourceable and are deliberately not attempted here)"
    ))
    t.font = Font(name="Arial", size=12, bold=True)

    # ==== Section 1: Injury History per current Starter/Backup QB =========================
    sec1_first = 5
    sec1_last = sec1_first + len(current_qbs) - 1
    _section_title(
        ws, 3, 7,
        "Section 1 \u2014 Injury History (this season's Starters + Backups; 3-year "
        "Out/Doubtful week count via nflverse weekly injury reports)",
    )
    _header_row(ws, 4, [
        "Player Name", "Player ID", "Team", "Role", "Out/Doubtful\nWeeks (3-Yr)",
        "Most Recent\nOut (Season)", "Most Recent\nOut (Week)",
    ])
    for i, qbrow in enumerate(current_qbs.to_dict("records")):
        row = sec1_first + i
        hist = injury_lookup.get(qbrow["Player ID"], {})
        values = [
            qbrow["Player Name"], qbrow["Player ID"], qbrow["Team"], qbrow["Role"],
            int(hist.get("Out/Doubtful Weeks (3-Yr)", 0)),
            hist.get("Most Recent Out Season"), hist.get("Most Recent Out Week"),
        ]
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = INPUT_FONT

    # ==== Section 2: Team Game Log (raw) + trailing-window Schedule Density ================
    sec2_title_row = sec1_last + 2
    sec2_header_row = sec2_title_row + 1
    sec2_first = sec2_header_row + 1
    sec2_last = sec2_first + len(game_log) - 1
    _section_title(
        ws, sec2_title_row, 5,
        "Section 2 \u2014 Team Game Log, Real 2025 Regular Season (raw material for the "
        "trailing-window formulas in Section 2b below; miles traveled via real stadium "
        "coordinates, 0 for home games)",
    )
    _header_row(
        ws, sec2_header_row,
        ["Team", "Season", "Week", "Game Date", "Miles Traveled\n(This Game)"], height=20,
    )
    for i, r in enumerate(game_log.to_dict("records")):
        row = sec2_first + i
        values = [
            r["Team"], int(r["Season"]), int(r["Week"]), r["Game Date"],
            float(r["Miles Traveled (This Game)"]),
        ]
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = INPUT_FONT
            if col == 4:
                cell.number_format = "yyyy-mm-dd"
            if col == 5:
                cell.number_format = "0"

    date_range = f"$D${sec2_first}:$D${sec2_last}"
    team_range = f"$A${sec2_first}:$A${sec2_last}"
    miles_range = f"$E${sec2_first}:$E${sec2_last}"
    as_of_date = game_log["Game Date"].max()

    sec2b_title_row = sec2_last + 2
    sec2b_header_row = sec2b_title_row + 1
    team_order = sorted(current_qbs["Team"].unique())
    sec2b_first = sec2b_header_row + 1
    sec2b_last = sec2b_first + len(team_order) - 1
    _section_title(
        ws, sec2b_title_row, 4,
        "Section 2b \u2014 Rolling Schedule Density, as of the last pulled game date "
        f"({as_of_date.date()}). For an upcoming season opener (Week 1) every team's "
        "trailing window is empty by definition -- this demonstrates the mechanism against "
        "real recent-season data, not a Week 1 snapshot.",
    )
    _header_row(ws, sec2b_header_row, [
        "Team", "Games in Trailing\n7 Days", "Games in Trailing\n14 Days",
        "Miles Traveled,\nTrailing 14 Days",
    ], height=20)
    for i, team in enumerate(team_order):
        row = sec2b_first + i
        ws.cell(row=row, column=1, value=team).font = INPUT_FONT
        g7 = ws.cell(row=row, column=2, value=(
            f'=COUNTIFS({team_range},$A{row},{date_range},">="&(DATE({as_of_date.year},'
            f'{as_of_date.month},{as_of_date.day})-7),{date_range},"<="&DATE('
            f'{as_of_date.year},{as_of_date.month},{as_of_date.day}))'
        ))
        g14 = ws.cell(row=row, column=3, value=(
            f'=COUNTIFS({team_range},$A{row},{date_range},">="&(DATE({as_of_date.year},'
            f'{as_of_date.month},{as_of_date.day})-14),{date_range},"<="&DATE('
            f'{as_of_date.year},{as_of_date.month},{as_of_date.day}))'
        ))
        m14 = ws.cell(row=row, column=4, value=(
            f'=SUMIFS({miles_range},{team_range},$A{row},{date_range},">="&(DATE('
            f'{as_of_date.year},{as_of_date.month},{as_of_date.day})-14),{date_range},'
            f'"<="&DATE({as_of_date.year},{as_of_date.month},{as_of_date.day}))'
        ))
        for cell in (g7, g14, m14):
            cell.font = FORMULA_FONT

    # ==== Section 3: Z-score + weighted Availability Score, per QB ========================
    sec3_title_row = sec2b_last + 2
    sec3_header_row = sec3_title_row + 1
    sec3_first = sec3_header_row + 1
    sec3_last = sec3_first + len(current_qbs) - 1

    league_avg_row = sec3_last + 2
    league_std_row = league_avg_row + 1

    _section_title(
        ws, sec3_title_row, 8,
        "Section 3 \u2014 Availability Score (Z-score each dimension, sign-flipped so "
        "higher Z always means MORE available/healthier, then combine via the new "
        "Availability Index weights on Model Assumptions)",
    )
    _header_row(ws, sec3_header_row, [
        "Player Name", "Team", "Injury History\n(Weeks, 3-Yr)", "Miles Traveled,\nTrailing 14 Days",
        "Injury\nZ (flipped)", "Schedule\nZ (flipped)", "Availability\nScore (pts)",
    ])

    for i in range(len(current_qbs)):
        row = sec3_first + i
        s1_row = sec1_first + i
        ws.cell(row=row, column=1, value=f"=A{s1_row}").font = FORMULA_FONT
        ws.cell(row=row, column=2, value=f"=C{s1_row}").font = FORMULA_FONT
        injury_cell = ws.cell(row=row, column=3, value=f"=E{s1_row}")
        injury_cell.font = FORMULA_FONT

        sched_cell = ws.cell(row=row, column=4, value=(
            f'=IFERROR(INDEX($D${sec2b_first}:$D${sec2b_last},'
            f'MATCH(B{row},$A${sec2b_first}:$A${sec2b_last},0)),0)'
        ))
        sched_cell.font = FORMULA_FONT

        iz = ws.cell(row=row, column=5, value=(
            f"=($C${league_avg_row}-C{row})/$C${league_std_row}"
        ))
        sz = ws.cell(row=row, column=6, value=(
            f"=($D${league_avg_row}-D{row})/$D${league_std_row}"
        ))
        iz.font = FORMULA_FONT
        sz.font = FORMULA_FONT
        iz.number_format = "0.00"
        sz.number_format = "0.00"

        score = ws.cell(row=row, column=7, value=(
            f"=E{row}*'Model Assumptions'!$C$41+F{row}*'Model Assumptions'!$C$42"
        ))
        score.font = FORMULA_FONT
        score.number_format = "0.0;(0.0)"

    ws.cell(row=league_avg_row, column=1, value="League Average").font = FORMULA_FONT
    ws.cell(row=league_std_row, column=1, value="League Std. Dev.").font = FORMULA_FONT
    injury_col_range = f"C${sec3_first}:C${sec3_last}"
    sched_col_range = f"D${sec3_first}:D${sec3_last}"
    avg_i = ws.cell(row=league_avg_row, column=3, value=f"=AVERAGE({injury_col_range})")
    avg_s = ws.cell(row=league_avg_row, column=4, value=f"=AVERAGE({sched_col_range})")
    std_i = ws.cell(row=league_std_row, column=3, value=f"=STDEVP({injury_col_range})")
    std_s = ws.cell(row=league_std_row, column=4, value=f"=STDEVP({sched_col_range})")
    for cell in (avg_i, avg_s, std_i, std_s):
        cell.font = FORMULA_FONT
        cell.number_format = "0.00"

    note_row = league_std_row + 2
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=8)
    note = ws.cell(row=note_row, column=1, value=(
        "PARTIAL index by design -- Section header and this note exist specifically so "
        "nobody mistakes this for the fuller 5-dimension sports-science version the user "
        "originally asked about (ACWR, GPS/internal-external load, and subjective wellness "
        "surveys are proprietary/internal-only data, never publicly released; see the spec "
        "for why). Section 1 (Injury History) is real: weeks listed Out or Doubtful on the "
        "public weekly injury report, which does not cleanly separate injury from coach's-"
        "decision in all cases, so treat it as a reasonable proxy, not a clinical record. "
        "Section 2/2b (Schedule Density) is genuinely computed from real game dates and "
        "real stadium-distance travel miles, but for an upcoming season-opener Week (like "
        "the current Week 1 Matchups tab), every team's trailing 7/14-day window is empty "
        "by definition -- there are no games yet this season to look back on. Section 2b is "
        "evaluated as of the last real pulled game date instead, to demonstrate the rolling "
        "mechanism against real data; re-run this script as the season progresses and it "
        "will reflect that week's actual trailing window automatically. Applied to this "
        "season's Starters + Backups only, per the spec -- extending to the full 53-man "
        "roster is a bigger lift, noted as a future step."
    ))
    note.font = NOTE_FONT
    note.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(workbook_path)
    print(f"Built '{SHEET_NAME}': {len(current_qbs)} QBs scored, {len(game_log)} game-log rows.")
    print(f"Saved to {workbook_path}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print('Usage: uv run python scripts/build_availability_index.py "path/to/workbook.xlsx"')
        sys.exit(1)
    build(sys.argv[1])
