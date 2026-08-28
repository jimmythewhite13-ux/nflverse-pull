"""
Builds Part C of claude_code_spec_qb_index.md: a new "QB Index" workbook tab, structurally
mirroring "Advanced Efficiency Metrics" (5-section decay-weighted/regressed pattern), plus
the new "QB Index Weighting & Conversion" assumptions on "Model Assumptions".

Section 1: raw 3-year data per QB-season (from qb_stats.py), Role-tagged
Section 2: league average per metric per season (Starters+Backups only) + a Rookie Baseline
           row (average of every qualifying rookie season, across all pulled years)
Section 3: per-QB (this season's Starters + Backups only) Y-1/Y-2/Y-3 -> decay-weighted
           average -> last-year-emphasis blend -> regressed toward league mean, per metric
           -- missing Y-2/Y-3 slots are substituted with the Rookie Baseline, not zero-
           filled or omitted, with a live "Years of Real History" (0-3) count per QB
Section 4: league average/std-dev of Section 3's Projected 3-Yr Baseline, per metric
Section 5: Z-score each metric (no sign-flip -- all 3 QB metrics are "higher is better"),
           weighted composite, and the final points-scale QB Index Score (50 + Z*10)

Usage:
    uv run python scripts/build_qb_index.py "C:\\path\\to\\NFL_Prediction_Model.xlsx"
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from nflverse_pull.efficiency import fetch_pbp  # noqa: E402
from nflverse_pull.qb_stats import compute_qb_roles, compute_team_season_qb_stats  # noqa: E402

YEARS = [2023, 2024, 2025]
SHEET_NAME = "QB Index"

METRICS = [
    {"key": "epa", "label": "EPA/Play", "sec1_col": "E", "fmt": "0.000"},
    {"key": "cpoe", "label": "CPOE", "sec1_col": "F", "fmt": "0.00"},
    {"key": "anya", "label": "ANY/A", "sec1_col": "G", "fmt": "0.00"},
]

# --- Styles (same conventions as build_efficiency_engine.py / the rest of the workbook) --
TITLE_FONT = Font(name="Arial", size=10, bold=True)
TITLE_FILL = PatternFill("solid", fgColor="FFD9E1F2")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="FF1F4E78")
HEADER_ALIGN = Alignment(wrap_text=True, horizontal="center", vertical="center")
INPUT_FONT = Font(name="Arial", size=10, color="FF0000FF")
FORMULA_FONT = Font(name="Arial", size=10, color="FF000000")
NOTE_FONT = Font(name="Arial", size=9, color="FF808080")
ASSUMPTION_FILL = PatternFill("solid", fgColor="FFFFFF00")


def _section_title(ws, row: int, last_col: int, text: str) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL


def _header_row(ws, row: int, headers: list[str], height: float = 28) -> None:
    ws.row_dimensions[row].height = height
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN


def add_model_assumptions_weights(wb: openpyxl.Workbook) -> None:
    ws = wb["Model Assumptions"]
    ws.merge_cells("A33:D33")
    title = ws.cell(row=33, column=1, value=(
        "QB Index Weighting & Conversion (pts per std. dev.; points-scale constants -- see "
        "'QB Index' tab)"
    ))
    title.font = HEADER_FONT
    title.fill = HEADER_FILL

    rows = [
        (34, "EPA/Play Weight (QB Index, pts per SD)", 0.5,
         "EPA/play is the most complete single stat for a QB too -- weighted highest, same "
         "logic as the team-level Advanced Efficiency Metrics weighting."),
        (35, "CPOE Weight (QB Index, pts per SD)", 0.3,
         "Completion % Over Expected isolates accuracy skill and is less scheme-dependent "
         "than EPA, but overlaps with it -- weighted lower to avoid double-counting."),
        (36, "ANY/A Weight (QB Index, pts per SD)", 0.3,
         "Traditional, well-understood counting stat; correlates with EPA but serves as a "
         "useful sanity check against it."),
        (37, "QB Index Baseline Score (league-average QB)", 50,
         "The points-scale center: a QB scoring exactly average across all three weighted "
         "metrics gets this score. Change this to re-center the whole scale."),
        (38, "QB Index Points per Std. Dev.", 10,
         "Converts the weighted Z-score sum into points: each 10 points above/below the "
         "baseline represents roughly 1 standard deviation from league average."),
        (39, "QB Index Points-to-Game-Points Conversion", 0.15,
         "A starting guess, like every other coefficient in this model: predicted scoring "
         "impact (in game points) per 1 point of QB Index gap between a team's starter and "
         "backup. Tune this if replacement-value swings feel too large or too small."),
    ]
    for row, label, value, note in rows:
        ws.cell(row=row, column=2, value=label)
        c = ws.cell(row=row, column=3, value=value)
        c.font = INPUT_FONT
        c.fill = ASSUMPTION_FILL
        c.number_format = "0.00"
        n = ws.cell(row=row, column=4, value=note)
        n.font = NOTE_FONT
        n.alignment = Alignment(wrap_text=True, vertical="top")


def build(workbook_path: str) -> dict:
    print(f"Pulling {YEARS} play-by-play data...")
    pbp = fetch_pbp(YEARS)
    stats = compute_team_season_qb_stats(pbp)
    roles = compute_qb_roles(stats)

    role_lookup = {
        (r["Player ID"], r["Season"]): r["Role"] for r in roles.to_dict("records")
    }
    stats = stats.copy()
    stats["Role"] = [
        role_lookup.get((pid, season), "Other")
        for pid, season in zip(stats["Player ID"], stats["Season"], strict=True)
    ]
    stats = stats.sort_values(["Team", "Season", "Dropbacks"], ascending=[True, True, False])
    stats = stats.reset_index(drop=True)

    current_season = max(YEARS)
    current_qbs = (
        roles[(roles["Season"] == current_season) & (roles["Role"].isin(["Starter", "Backup"]))]
        .sort_values(["Team", "Role"])
        .reset_index(drop=True)
    )
    n_qb = len(current_qbs)
    print(f"{len(stats)} QB-seasons in Section 1; {n_qb} current Starters/Backups scored.")

    wb = openpyxl.load_workbook(workbook_path)
    add_model_assumptions_weights(wb)

    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME, index=wb.sheetnames.index("Advanced Efficiency Metrics") + 1)
    ws.column_dimensions["A"].width = 18.0
    ws.column_dimensions["C"].width = 20.0

    ws.merge_cells("A1:I1")
    t = ws.cell(row=1, column=1, value=(
        "QB Index -- Multi-Year Decay-Weighted QB Rating (EPA/Play, CPOE, ANY/A), mirrors "
        "Advanced Efficiency Metrics / YoY Baseline Engine"
    ))
    t.font = Font(name="Arial", size=12, bold=True)

    # ==== Section 1: Raw 3-year data per QB-season ========================================
    sec1_first_row = 5
    sec1_last_row = sec1_first_row + len(stats) - 1
    _section_title(ws, 3, 9, "Section 1 \u2014 Raw 3-Year Data per QB-Season (from nflverse pbp)")
    _header_row(
        ws, 4,
        ["Player Name", "Player ID", "Team", "Season", "EPA/Play", "CPOE", "ANY/A",
         "Is Rookie Season", "Role"],
    )
    for i, r in enumerate(stats.to_dict("records")):
        row = sec1_first_row + i
        values = [
            r["Player Name"], r["Player ID"], r["Team"], int(r["Season"]),
            float(r["EPA/Play"]), float(r["CPOE"]), float(r["ANY/A"]),
            bool(r["Is Rookie Season"]), r["Role"],
        ]
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = INPUT_FONT
            if col == 5:
                cell.number_format = "0.000"
            elif col in (6, 7):
                cell.number_format = "0.00"

    id_range = f"$B${sec1_first_row}:$B${sec1_last_row}"
    season_range = f"$D${sec1_first_row}:$D${sec1_last_row}"
    rookie_range = f"$H${sec1_first_row}:$H${sec1_last_row}"
    role_range = f"$I${sec1_first_row}:$I${sec1_last_row}"
    metric_ranges = {
        m["key"]: f"${m['sec1_col']}${sec1_first_row}:${m['sec1_col']}${sec1_last_row}"
        for m in METRICS
    }

    # ==== Section 2: League average per season (Starters+Backups only) + Rookie Baseline ==
    sec2_title_row = sec1_last_row + 2
    sec2_header_row = sec2_title_row + 1
    season_rows = {yr: sec2_header_row + 1 + i for i, yr in enumerate(YEARS)}
    rookie_baseline_row = sec2_header_row + 1 + len(YEARS)
    rookie_count_row = rookie_baseline_row + 1

    _section_title(
        ws, sec2_title_row, 4,
        "Section 2 \u2014 League Average per Season (Starters + Backups only) and Rookie "
        "Baseline (all qualifying rookie seasons, any role, across every pulled year)",
    )
    _header_row(ws, sec2_header_row, ["Season / Stat", *[m["label"] for m in METRICS]], height=20)

    for yr in YEARS:
        row = season_rows[yr]
        c = ws.cell(row=row, column=1, value=yr)
        c.font = INPUT_FONT
        for j, m in enumerate(METRICS):
            formula = (
                f"=AVERAGEIFS({metric_ranges[m['key']]},{season_range},{yr},"
                f"{role_range},\"<>Other\")"
            )
            cell = ws.cell(row=row, column=2 + j, value=formula)
            cell.font = FORMULA_FONT
            cell.number_format = m["fmt"]

    lbl = ws.cell(row=rookie_baseline_row, column=1, value="Rookie Baseline (all years)")
    lbl.font = FORMULA_FONT
    for j, m in enumerate(METRICS):
        formula = f"=AVERAGEIF({rookie_range},TRUE,{metric_ranges[m['key']]})"
        cell = ws.cell(row=rookie_baseline_row, column=2 + j, value=formula)
        cell.font = FORMULA_FONT
        cell.number_format = m["fmt"]

    lbl2 = ws.cell(
        row=rookie_count_row, column=1,
        value="Rookie Baseline Sample Size (qualifying rookie seasons) -- small-sample "
              "estimate with only 3 years pulled; recomputes automatically as more seasons "
              "are added",
    )
    lbl2.font = NOTE_FONT
    lbl2.alignment = Alignment(wrap_text=True)
    for j in range(len(METRICS)):
        cell = ws.cell(row=rookie_count_row, column=2 + j, value=f"=COUNTIF({rookie_range},TRUE)")
        cell.font = FORMULA_FONT
        cell.number_format = "0"

    rookie_baseline_cell = {
        m["key"]: f"${get_column_letter(2 + j)}${rookie_baseline_row}"
        for j, m in enumerate(METRICS)
    }
    season_avg_col = {m["key"]: get_column_letter(2 + j) for j, m in enumerate(METRICS)}

    # ==== Section 3: Per-QB decay-weighted, regressed baseline, per metric ================
    sec3_title_row = rookie_count_row + 2
    sec3_header_row = sec3_title_row + 1
    sec3_first_row = sec3_header_row + 1
    sec3_last_row = sec3_first_row + n_qb - 1
    sec3_last_col = 5 + len(METRICS) * 7  # Player/ID/Team/Role/YearsHistory + 7 cols x 3

    _section_title(
        ws, sec3_title_row, sec3_last_col,
        "Section 3 \u2014 Per-QB 3-Yr Decay-Weighted, Regressed Baseline (this season's "
        "Starters + Backups only; missing Y-2/Y-3 seasons substitute the Rookie Baseline "
        "above, not zero -- same formula chain as Advanced Efficiency Metrics Section 3)",
    )
    headers = ["Player Name", "Player ID", "Team", "Role", "Years of\nReal History"]
    metric_block_start_col: dict[str, int] = {}
    col_cursor = 6
    for m in METRICS:
        metric_block_start_col[m["key"]] = col_cursor
        headers += [
            f"{m['label']} Y-1", f"{m['label']} Y-2", f"{m['label']} Y-3",
            f"Weighted\n{m['label']} Avg\n(3-Yr decay)", f"Team History\n{m['label']}",
            f"League Baseline\n{m['label']} (Y-1)", f"Projected 3-Yr\n{m['label']} Baseline",
        ]
        col_cursor += 7
    _header_row(ws, sec3_header_row, headers)

    for i, qb in enumerate(current_qbs.to_dict("records")):
        row = sec3_first_row + i
        ws.cell(row=row, column=1, value=qb["Player Name"]).font = FORMULA_FONT
        ws.cell(row=row, column=2, value=qb["Player ID"]).font = FORMULA_FONT
        ws.cell(row=row, column=3, value=qb["Team"]).font = FORMULA_FONT
        ws.cell(row=row, column=4, value=qb["Role"]).font = FORMULA_FONT

        years_hist_terms = "+".join(
            f"--(COUNTIFS({id_range},$B{row},{season_range},'Model Assumptions'!$C$18-{k})>0)"
            for k in (1, 2, 3)
        )
        yh = ws.cell(row=row, column=5, value=f"={years_hist_terms}")
        yh.font = FORMULA_FONT
        yh.number_format = "0"

        for m in METRICS:
            base = metric_block_start_col[m["key"]]
            y1, y2, y3, wavg, th, lb, pb = (get_column_letter(base + k) for k in range(7))
            mrange = metric_ranges[m["key"]]
            rbcell = rookie_baseline_cell[m["key"]]

            def _ysub(offset: int) -> str:
                return (
                    f"=IF(COUNTIFS({id_range},$B{row},{season_range},"
                    f"'Model Assumptions'!$C$18-{offset})=0,{rbcell},"
                    f"SUMIFS({mrange},{id_range},$B{row},{season_range},"
                    f"'Model Assumptions'!$C$18-{offset}))"
                )

            f_y1 = ws.cell(row=row, column=base, value=_ysub(1))
            f_y2 = ws.cell(row=row, column=base + 1, value=_ysub(2))
            f_y3 = ws.cell(row=row, column=base + 2, value=_ysub(3))
            f_wavg = ws.cell(row=row, column=base + 3, value=(
                f"=({y1}{row}*1+{y2}{row}*'Model Assumptions'!$C$20+"
                f"{y3}{row}*('Model Assumptions'!$C$20^2))/"
                f"(1+'Model Assumptions'!$C$20+'Model Assumptions'!$C$20^2)"
            ))
            f_th = ws.cell(row=row, column=base + 4, value=(
                f"={y1}{row}*'Model Assumptions'!$C$22+{wavg}{row}*(1-'Model Assumptions'!$C$22)"
            ))
            sec2_col = season_avg_col[m["key"]]
            f_lb = ws.cell(row=row, column=base + 5, value=(
                f"=INDEX(${sec2_col}${season_rows[YEARS[0]]}:${sec2_col}${season_rows[YEARS[-1]]},"
                f"MATCH('Model Assumptions'!$C$18-1,$A${season_rows[YEARS[0]]}:"
                f"$A${season_rows[YEARS[-1]]},0))"
            ))
            f_pb = ws.cell(row=row, column=base + 6, value=(
                f"={th}{row}*'Model Assumptions'!$C$21+{lb}{row}*(1-'Model Assumptions'!$C$21)"
            ))
            for cell in (f_y1, f_y2, f_y3, f_wavg, f_th, f_lb, f_pb):
                cell.font = FORMULA_FONT
                cell.number_format = m["fmt"]

    # ==== Section 4: League average/std-dev of Section 3's Projected Baselines ============
    sec4_title_row = sec3_last_row + 2
    sec4_header_row = sec4_title_row + 1
    avg_row, std_row = sec4_header_row + 1, sec4_header_row + 2

    _section_title(
        ws, sec4_title_row, 4, "Section 4 \u2014 League Average & Std. Dev. of the 3-Yr Baselines"
    )
    _header_row(ws, sec4_header_row, ["Stat", *[m["label"] for m in METRICS]], height=18)

    ws.cell(row=avg_row, column=1, value="League Average").font = FORMULA_FONT
    ws.cell(row=std_row, column=1, value="League Std. Dev.").font = FORMULA_FONT

    proj_baseline_col = {
        m["key"]: get_column_letter(metric_block_start_col[m["key"]] + 6) for m in METRICS
    }
    for j, m in enumerate(METRICS):
        pcol = proj_baseline_col[m["key"]]
        pb_range = f"{pcol}${sec3_first_row}:{pcol}${sec3_last_row}"
        a = ws.cell(row=avg_row, column=2 + j, value=f"=AVERAGE({pb_range})")
        s = ws.cell(row=std_row, column=2 + j, value=f"=STDEVP({pb_range})")
        for cell in (a, s):
            cell.font = FORMULA_FONT
            cell.number_format = m["fmt"]

    # ==== Section 5: Z-scores, weighted composite, points-scale QB Index Score ============
    sec5_title_row = std_row + 2
    sec5_header_row = sec5_title_row + 1
    sec5_first_row = sec5_header_row + 1
    sec5_last_row = sec5_first_row + n_qb - 1

    _section_title(
        ws, sec5_title_row, 10,
        "Section 5 \u2014 Z-Scores and QB Index Score (all three metrics are \"higher is "
        "better\" for a QB -- no sign-flip needed, unlike the team-level defense/allowed "
        "metrics)",
    )
    _header_row(
        ws, sec5_header_row,
        ["Player Name", "Player ID", "Team", "Role", *[f"{m['label']}\nZ" for m in METRICS],
         "Weighted\nZ-Score Sum", "QB Index\nScore (Points)", "Years of\nReal History"],
    )

    avg_cell_ref = {
        m["key"]: f"${get_column_letter(2 + j)}${avg_row}" for j, m in enumerate(METRICS)
    }
    std_cell_ref = {
        m["key"]: f"${get_column_letter(2 + j)}${std_row}" for j, m in enumerate(METRICS)
    }
    weight_cells = {"epa": "$C$34", "cpoe": "$C$35", "anya": "$C$36"}

    for i in range(n_qb):
        sec3_row = sec3_first_row + i
        row = sec5_first_row + i
        for col, src_col in ((1, "A"), (2, "B"), (3, "C"), (4, "D")):
            f = ws.cell(row=row, column=col, value=f"={src_col}{sec3_row}")
            f.font = FORMULA_FONT

        z_cols = []
        for j, m in enumerate(METRICS):
            pcol = proj_baseline_col[m["key"]]
            formula = f"=({pcol}{sec3_row}-{avg_cell_ref[m['key']]})/{std_cell_ref[m['key']]}"
            col = 5 + j
            cell = ws.cell(row=row, column=col, value=formula)
            cell.font = FORMULA_FONT
            cell.number_format = "0.00"
            z_cols.append(get_column_letter(col))

        weighted_terms = " + ".join(
            f"{z_cols[j]}{row}*'Model Assumptions'!{weight_cells[m['key']]}"
            for j, m in enumerate(METRICS)
        )
        wz_col = 5 + len(METRICS)
        wz = ws.cell(row=row, column=wz_col, value=f"={weighted_terms}")
        wz.font = FORMULA_FONT
        wz.number_format = "0.00"

        score = ws.cell(row=row, column=wz_col + 1, value=(
            f"='Model Assumptions'!$C$37+{get_column_letter(wz_col)}{row}*'Model Assumptions'!$C$38"
        ))
        score.font = FORMULA_FONT
        score.number_format = "0.0;(0.0)"

        yh_link = ws.cell(row=row, column=wz_col + 2, value=f"=E{sec3_row}")
        yh_link.font = FORMULA_FONT
        yh_link.number_format = "0"

    # ---- Closing note --------------------------------------------------------------------
    note_row = sec5_last_row + 2
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=10)
    note = ws.cell(row=note_row, column=1, value=(
        "Mirrors Advanced Efficiency Metrics' 5-section decay-weighted/regressed pattern, "
        "run per QB instead of per team. Section 1 is the raw per-QB-season table from "
        "nflverse_pull.qb_stats (EPA/Play, CPOE, ANY/A; QB-seasons under 100 dropbacks are "
        "excluded entirely, not zero-filled). Is Rookie Season is a proxy (first pbp-"
        "observed qualifying season within the pulled years, not real draft data) -- a "
        "veteran whose first pulled-window season happens to be the earliest year pulled "
        "will show as a false rookie; this is a documented simplification, not a bug. "
        "Section 2 computes each metric's league average among Starters+Backups per season, "
        "plus a Rookie Baseline (average of every qualifying rookie season, any role, across "
        "all pulled years) used to fill in a QB's missing Y-2/Y-3 history below rather than "
        "zero-filling or omitting it -- both the Section 2 averages and the Rookie Baseline "
        "are formulas over Section 1, so they recompute automatically as more data is "
        "pulled. Section 3 scores only this season's Starters and Backups (from Model "
        "Assumptions Current Season - 1); its Years of Real History column (0-3) shows at a "
        "glance how much of a QB's rating leans on real career data versus the Rookie "
        "Baseline. Section 4/5 Z-score Section 3's output and convert to the QB Index Score "
        "(points-scale, 50 = league average, +/-10 per std. dev., both tunable on Model "
        "Assumptions) using the new EPA/CPOE/ANY-A weights there. One known limitation: a "
        "QB traded mid-season keeps two Section 1 rows (one per team, per Part A's own "
        "design) -- his Y-1/Y-2/Y-3 SUMIFS would sum both teams' per-play averages together "
        "for that one season rather than blend them correctly; rare in practice, but worth "
        "knowing if a current Starter/Backup was traded within the pulled window."
    ))
    note.font = NOTE_FONT
    note.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(workbook_path)
    print(f"Built '{SHEET_NAME}': Section 1 {len(stats)} rows, Section 3/5 {n_qb} QBs.")
    print(f"Saved to {workbook_path}")
    return {
        "sec1_rows": len(stats), "n_qb": n_qb,
        "sec3_range": (sec3_first_row, sec3_last_row),
        "sec5_range": (sec5_first_row, sec5_last_row),
    }


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print('Usage: uv run python scripts/build_qb_index.py "path/to/workbook.xlsx"')
        sys.exit(1)
    build(sys.argv[1])
