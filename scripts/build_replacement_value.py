"""
Builds Part D of claude_code_spec_qb_index.md: the Replacement Value Index (a new Section 6
on the "QB Index" tab), then wires it into "Team Ratings" and "Week 1 Matchups" as a live,
recalculating point adjustment -- not a snapshot.

Requires "QB Index" to already exist (built by scripts/build_qb_index.py first).

Usage:
    uv run python scripts/build_replacement_value.py "C:\\path\\to\\NFL_Prediction_Model.xlsx"
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

# Same canonical 32-team row order as build_efficiency_engine.py (copied, not imported --
# that script lives on a separate, still-unmerged branch/PR at the time of writing; see
# 'YoY Baseline Engine'!A110:A141, the original source of this order).
TEAM_ORDER = [
    "Buffalo Bills", "Miami Dolphins", "New England Patriots", "New York Jets",
    "Baltimore Ravens", "Cincinnati Bengals", "Cleveland Browns", "Pittsburgh Steelers",
    "Houston Texans", "Indianapolis Colts", "Jacksonville Jaguars", "Tennessee Titans",
    "Denver Broncos", "Kansas City Chiefs", "Las Vegas Raiders", "Los Angeles Chargers",
    "Dallas Cowboys", "New York Giants", "Philadelphia Eagles", "Washington Commanders",
    "Chicago Bears", "Detroit Lions", "Green Bay Packers", "Minnesota Vikings",
    "Atlanta Falcons", "Carolina Panthers", "New Orleans Saints", "Tampa Bay Buccaneers",
    "Arizona Cardinals", "Los Angeles Rams", "San Francisco 49ers", "Seattle Seahawks",
]

QB_INDEX_SHEET = "QB Index"
TEAM_RATINGS_SHEET = "Team Ratings"
MATCHUPS_SHEET = "Week 1 Matchups"

TITLE_FONT = Font(name="Arial", size=10, bold=True)
TITLE_FILL = PatternFill("solid", fgColor="FFD9E1F2")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="FF1F4E78")
HEADER_ALIGN = Alignment(wrap_text=True, horizontal="center", vertical="center")
INPUT_FONT = Font(name="Arial", size=10, color="FF0000FF")
FORMULA_FONT = Font(name="Arial", size=10, color="FF000000")
LINK_FONT = Font(name="Arial", size=10, color="FF008000")
NOTE_FONT = Font(name="Arial", size=9, color="FF808080")


def append_term_once(formula: str, term: str) -> str:
    """
    Pure function. Appends `term` (e.g. "+AS3") to the end of `formula` exactly once,
    stripping any number of pre-existing trailing copies first. Makes re-running this
    script against a workbook it already touched idempotent instead of re-appending and
    silently multiplying the term on every run -- this bit the ongoing automated pipeline
    for real (see commit history): three manual test runs left Z3 reading
    "...+AS3+AS3+AS3" before this existed.
    """
    escaped = re.escape(term)
    stripped = re.sub(rf"({escaped})+$", "", formula)
    return f"{stripped}{term}"


def _section_title(ws, row: int, last_col: int, text: str) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL


def _header_row(ws, row: int, headers: list[str], height: float = 28) -> None:
    ws.row_dimensions[row].height = height
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN


def _find_section5_range(ws) -> tuple[int, int]:
    """Scans column A for the 'Section 5' title, then finds its header + data rows."""
    title_row = None
    for row in range(1, ws.max_row + 1):
        v = ws.cell(row=row, column=1).value
        if isinstance(v, str) and v.startswith("Section 5"):
            title_row = row
            break
    if title_row is None:
        raise ValueError(
            "Could not find 'Section 5' on the QB Index tab -- run build_qb_index.py first."
        )
    header_row = title_row + 1
    first_data_row = header_row + 1
    row = first_data_row
    while ws.cell(row=row, column=1).value is not None:
        row += 1
    return first_data_row, row - 1


def build(workbook_path: str) -> None:
    wb = openpyxl.load_workbook(workbook_path)
    qb = wb[QB_INDEX_SHEET]

    sec5_first, sec5_last = _find_section5_range(qb)
    print(f"Section 5 data: rows {sec5_first}-{sec5_last}")

    # ==== Section 6: Replacement Value Index (new, on QB Index tab) =======================
    sec6_title_row = sec5_last + 2
    sec6_header_row = sec6_title_row + 1
    sec6_first_row = sec6_header_row + 1
    sec6_last_row = sec6_first_row + len(TEAM_ORDER) - 1

    _section_title(
        qb, sec6_title_row, 7,
        "Section 6 \u2014 Replacement Value Index (Starter QB Index Score minus Backup QB "
        "Index Score, per team; can be negative -- a negative value means the backup rates "
        "HIGHER than the starter, and the game-point adjustment below flips sign "
        "accordingly rather than assuming a backup is always a downgrade)",
    )
    _header_row(qb, sec6_header_row, [
        "Team", "Starter Name", "Starter QB\nIndex Score", "Backup Name", "Backup QB\nIndex Score",
        "Replacement Value\n(Index Points)", "Replacement Value\n(Game Points)",
    ])

    score_range = f"$I${sec5_first}:$I${sec5_last}"
    key_range = f"$K${sec5_first}:$K${sec5_last}"  # helper key column, written below
    name_range = f"$A${sec5_first}:$A${sec5_last}"

    # Helper key (Team|Role) in Section 5's column K, so a single MATCH can find a team's
    # specific Starter or Backup row -- MATCH can't take two criteria directly.
    for r in range(sec5_first, sec5_last + 1):
        qb.cell(row=r, column=11, value=f'=C{r}&"|"&D{r}').font = FORMULA_FONT

    for i, team in enumerate(TEAM_ORDER):
        row = sec6_first_row + i
        t = qb.cell(row=row, column=1, value=team)
        t.font = INPUT_FONT

        starter_key = f'"{team}|Starter"'
        backup_key = f'"{team}|Backup"'

        starter_name = qb.cell(row=row, column=2, value=(
            f"=IFERROR(INDEX({name_range},MATCH({starter_key},{key_range},0)),\"\")"
        ))
        starter_score = qb.cell(row=row, column=3, value=(
            f"=IFERROR(INDEX({score_range},MATCH({starter_key},{key_range},0)),\"\")"
        ))
        backup_name = qb.cell(row=row, column=4, value=(
            f"=IFERROR(INDEX({name_range},MATCH({backup_key},{key_range},0)),\"\")"
        ))
        backup_score = qb.cell(row=row, column=5, value=(
            f"=IFERROR(INDEX({score_range},MATCH({backup_key},{key_range},0)),\"\")"
        ))
        for cell in (starter_name, starter_score, backup_name, backup_score):
            cell.font = FORMULA_FONT
        starter_score.number_format = "0.0;(0.0)"
        backup_score.number_format = "0.0;(0.0)"

        rv_index = qb.cell(row=row, column=6, value=(
            f'=IF(OR(C{row}="",E{row}=""),"",C{row}-E{row})'
        ))
        rv_game = qb.cell(row=row, column=7, value=(
            f'=IF(F{row}="","",F{row}*\'Model Assumptions\'!$C$39)'
        ))
        rv_index.font = FORMULA_FONT
        rv_game.font = FORMULA_FONT
        rv_index.number_format = "0.0;(0.0)"
        rv_game.number_format = "0.00;(0.00)"

    note_row = sec6_last_row + 2
    qb.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=10)
    note = qb.cell(row=note_row, column=1, value=(
        "Replacement Value (Index Points) = Starter QB Index Score - Backup QB Index Score, "
        "per team; Replacement Value (Game Points) applies the "
        "Points-to-Game-Points Conversion (Model Assumptions C39). Both recalculate live "
        "from Section 5 -- there is nothing cached or hardcoded here. A team whose Starter "
        "or Backup never reached the 100-dropback qualifying threshold (see qb_stats.py) "
        "shows blank rather than a misleading 0, since no meaningful gap can be computed "
        "without both scores. 'Team Ratings' and 'Week 1 Matchups' each pull this column "
        "(G) directly, gated by their own 'Starting QB Status' input -- see those tabs."
    ))
    note.font = NOTE_FONT
    note.alignment = Alignment(wrap_text=True, vertical="top")

    # ==== Wire into Team Ratings ============================================================
    tr = wb[TEAM_RATINGS_SHEET]
    tr.cell(row=2, column=15, value="Starting QB\nStatus").font = HEADER_FONT
    tr.cell(row=2, column=15).fill = HEADER_FILL
    tr.cell(row=2, column=15).alignment = HEADER_ALIGN
    tr.cell(row=2, column=16, value="QB Replacement\nValue Adj (pts)").font = HEADER_FONT
    tr.cell(row=2, column=16).fill = HEADER_FILL
    tr.cell(row=2, column=16).alignment = HEADER_ALIGN

    rv_game_range = f"'{QB_INDEX_SHEET}'!$G${sec6_first_row}:$G${sec6_last_row}"
    rv_team_range = f"'{QB_INDEX_SHEET}'!$A${sec6_first_row}:$A${sec6_last_row}"

    for row in range(3, 3 + len(TEAM_ORDER)):
        status = tr.cell(row=row, column=15, value="Starter In")
        status.font = INPUT_FONT
        adj = tr.cell(row=row, column=16, value=(
            f'=IF(O{row}="Backup In",'
            f'-IFERROR(INDEX({rv_game_range},MATCH(A{row},{rv_team_range},0)),0),0)'
        ))
        adj.font = LINK_FONT
        adj.number_format = "0.00;(0.00)"

        # Net Power Rating (N) now also includes the QB Replacement Value adjustment. The
        # existing manual Injury/Replacement Adjustment (L) is untouched -- both apply.
        net = tr.cell(row=row, column=14, value=f"=J{row}-K{row}+L{row}+M{row}+P{row}")
        net.font = FORMULA_FONT
        net.number_format = "0.0;(0.0)"

    note_row_tr = 3 + len(TEAM_ORDER) + 1
    tr.merge_cells(start_row=note_row_tr, start_column=1, end_row=note_row_tr, end_column=16)
    tr_note = tr.cell(row=note_row_tr, column=1, value=(
        "Starting QB Status (O) is a manual per-team toggle: 'Starter In' (default, no "
        "adjustment) or 'Backup In'. When set to 'Backup In', QB Replacement Value Adj (P) "
        "automatically pulls -(that team's Replacement Value in game points) from 'QB "
        "Index' Section 6 -- if the backup actually rates HIGHER than the starter (a "
        "negative Replacement Value), this correctly becomes a POSITIVE adjustment, not a "
        "penalty. This is separate from the existing manual Injury / Replacement Adjustment "
        "(L), which is untouched -- both apply to Net Power Rating (N) if both are set."
    ))
    tr_note.font = NOTE_FONT
    tr_note.alignment = Alignment(wrap_text=True, vertical="top")

    # ==== Wire into Week 1 Matchups =========================================================
    wm = wb[MATCHUPS_SHEET]
    game_rows = []
    row = 3
    while wm.cell(row=row, column=1).value is not None:
        game_rows.append(row)
        row += 1

    headers = [
        (43, "Home Starting\nQB Status"), (44, "Away Starting\nQB Status"),
        (45, "Home QB\nReplacement Adj (pts)"), (46, "Away QB\nReplacement Adj (pts)"),
    ]
    for col, label in headers:
        c = wm.cell(row=2, column=col, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN

    for r in game_rows:
        home_status = wm.cell(row=r, column=43, value="Starter In")
        away_status = wm.cell(row=r, column=44, value="Starter In")
        home_status.font = INPUT_FONT
        away_status.font = INPUT_FONT

        home_adj = wm.cell(row=r, column=45, value=(
            f'=IF(AQ{r}="Backup In",'
            f'-IFERROR(INDEX({rv_game_range},MATCH(D{r},{rv_team_range},0)),0),0)'
        ))
        away_adj = wm.cell(row=r, column=46, value=(
            f'=IF(AR{r}="Backup In",'
            f'-IFERROR(INDEX({rv_game_range},MATCH(C{r},{rv_team_range},0)),0),0)'
        ))
        home_adj.font = LINK_FONT
        away_adj.font = LINK_FONT
        home_adj.number_format = "0.00;(0.00)"
        away_adj.number_format = "0.00;(0.00)"

        # Append the new adjustment to the existing Model Home/Away Score formulas (Z/AA),
        # at full weight -- same convention the existing V/W injury columns already use.
        # append_term_once() keeps this idempotent across repeated runs (see its docstring).
        z_cell = wm.cell(row=r, column=26)  # Z
        aa_cell = wm.cell(row=r, column=27)  # AA
        z_cell.value = append_term_once(z_cell.value, f"+AS{r}")
        aa_cell.value = append_term_once(aa_cell.value, f"+AT{r}")

    note_row_wm = max(game_rows) + 3
    wm.merge_cells(start_row=note_row_wm, start_column=1, end_row=note_row_wm, end_column=20)
    wm_note = wm.cell(row=note_row_wm, column=1, value=(
        "Home/Away Starting QB Status (AQ/AR) are per-game manual toggles -- set to 'Backup "
        "In' to automatically pull that team's QB Replacement Value into this week's Model "
        "Home/Away Score (Z/AA now include +AS/+AT respectively, same full-weight "
        "convention as the existing Home/Away Add'l Injury columns V/W). This is a live "
        "formula, not a snapshot: it reflects whatever Team Ratings / QB Index currently "
        "evaluate to, so changing a QB's current-season inputs or Model Assumptions "
        "reflects here automatically -- nothing needs re-entering per week."
    ))
    wm_note.font = NOTE_FONT
    wm_note.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(workbook_path)
    print(
        f"Built Section 6 ({len(TEAM_ORDER)} teams) on '{QB_INDEX_SHEET}', wired into "
        f"'{TEAM_RATINGS_SHEET}' (cols O/P) and '{MATCHUPS_SHEET}' "
        f"(cols AQ-AT, {len(game_rows)} games)."
    )
    print(f"Saved to {workbook_path}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print('Usage: uv run python scripts/build_replacement_value.py "path/to/workbook.xlsx"')
        sys.exit(1)
    build(sys.argv[1])
