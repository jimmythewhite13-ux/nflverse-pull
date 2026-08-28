"""
Builds Part 2 of claude_code_spec_current_roster_fix.md: a manual Starter/Backup override
input table on the "QB Index" tab. Blank = use the pulled current-roster data (Part 1).
Filled in = takes precedence -- real news (a beat-writer report, a surprise trade, a camp
competition outcome) moves faster than any pull job.

This table is standalone infrastructure only -- nothing reads it yet (Step 1 scope
explicitly excludes Part 4, the QB Index retrofit that will actually consume it).

Requires "QB Index" to already exist.

Usage:
    uv run python scripts/add_manual_override_table.py "C:\\path\\to\\NFL_Prediction_Model.xlsx"
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

TITLE_FONT = Font(name="Arial", size=10, bold=True)
TITLE_FILL = PatternFill("solid", fgColor="FFD9E1F2")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="FF1F4E78")
HEADER_ALIGN = Alignment(wrap_text=True, horizontal="center", vertical="center")
INPUT_FONT = Font(name="Arial", size=10, color="FF0000FF")
NOTE_FONT = Font(name="Arial", size=9, color="FF808080")

TEAM_ORDER = [
    "Buffalo Bills", "Miami Dolphins", "New England Patriots", "New York Jets",
    "Baltimore Ravens", "Cincinnati Bengals", "Cleveland Browns", "Pittsburgh Steelers",
    "Houston Texans", "Indianapolis Colts", "Jacksonville Jaguars", "Tennessee Titans",
    "Denver Broncos", "Kansas City Chiefs", "Las Vegas Raiders", "Los Angeles Chargers",
    "Dallas Cowboys", "New York Giants", "Philadelphia Eagles", "Washington Commanders",
    "Chicago Bears", "Detroit Lions", "Green Bay Packers", "Minnesota Vikings",
    "Atlanta Falcons", "Carolina Panthers", "New Orleans Saints", "Tampa Bay Buccaneers",
    "Arizona Cardinals", "Los Angeles Rams", "San Francisco 49ers", "Seattle Seahawks",
]

SHEET_NAME = "QB Index"
SECTION_MARKER = "Section 7 \u2014 Manual Roster Override"


def build(workbook_path: str) -> None:
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb[SHEET_NAME]

    # Idempotent: if this section already exists (a re-run), clear it and its data below
    # before rebuilding, rather than appending a second copy further down the sheet.
    existing_title_row = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.startswith(SECTION_MARKER):
            existing_title_row = r
            break
    if existing_title_row is not None:
        for r in range(existing_title_row, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c, value=None)
        title_row = existing_title_row
    else:
        # Find the first fully-blank row after all existing content, with a one-row gap.
        title_row = ws.max_row + 2

    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=4)
    t = ws.cell(row=title_row, column=1, value=(
        f"{SECTION_MARKER} (blank = use the pulled current-roster data; filled in = takes "
        "precedence over it -- a beat-writer report or camp-competition outcome will always "
        "be more current than the last pull. Not yet wired into Section 3/5/6 scoring -- "
        "see claude_code_spec_current_roster_fix.md Part 4.)"
    ))
    t.font = TITLE_FONT
    t.fill = TITLE_FILL

    header_row = title_row + 1
    headers = ["Team", "Manual Starter Override", "Manual Backup Override"]
    ws.row_dimensions[header_row].height = 20
    for col, label in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN

    first_data_row = header_row + 1
    for i, team in enumerate(TEAM_ORDER):
        row = first_data_row + i
        ws.cell(row=row, column=1, value=team).font = INPUT_FONT
        ws.cell(row=row, column=2, value=None).font = INPUT_FONT
        ws.cell(row=row, column=3, value=None).font = INPUT_FONT
    last_data_row = first_data_row + len(TEAM_ORDER) - 1

    wb.save(workbook_path)
    print(f"Built '{SECTION_MARKER}' on '{SHEET_NAME}': rows {first_data_row}-{last_data_row}.")
    print(f"Saved to {workbook_path}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print('Usage: uv run python scripts/add_manual_override_table.py "path/to/workbook.xlsx"')
        sys.exit(1)
    build(sys.argv[1])
