"""
Writes freshly pulled nflverse data into the NFL Prediction Model workbook's placeholder
input tables, replacing the SAMPLE PLACEHOLDER rows in place:

    'YoY Baseline Engine' Section 1          <- pull.py (per-team, per-season Off/Def PPG)
    'Advanced Efficiency Metrics' Section 1  <- efficiency.py (per-team EPA/Success Rate/NY/A)

Only those two input tables are touched -- every formula elsewhere in the workbook
(Section 2/3 on both tabs, Team Ratings, Week 1 Matchups, etc.) is left alone and
recalculates itself from the new inputs the next time the workbook is opened in Excel.

Run this whenever you want the workbook refreshed with the latest nflverse data:

    uv run python -m nflverse_pull.update_workbook "C:\\path\\to\\NFL_Prediction_Model.xlsx"
"""
from __future__ import annotations

import argparse

import openpyxl
import pandas as pd

from nflverse_pull.efficiency import compute_raw_efficiency, fetch_pbp
from nflverse_pull.pull import fetch_schedules, transform_to_team_season

YOY_SHEET = "YoY Baseline Engine"
EFFICIENCY_SHEET = "Advanced Efficiency Metrics"

# Column order matches efficiency.py's compute_raw_efficiency() output and the
# 'Advanced Efficiency Metrics' Section 1 headers (columns B-H).
EFFICIENCY_COLUMNS = [
    "epa_off", "epa_def", "success_off", "success_def", "nya_off", "nya_def", "proe",
]


def build_yoy_updates(ppg: pd.DataFrame) -> dict[tuple[str, int], tuple[float, float]]:
    """
    Pure function. Maps pull.transform_to_team_season()'s per-season output to
    {(Team, Season): (Off PPG, Def PPG)}, for writing into YoY Baseline Engine Section 1.
    """
    # Column names contain spaces ("Off PPG"), which itertuples/namedtuple would silently
    # rename to positional fields -- .to_dict("records") keeps the lookup explicit instead.
    return {
        (r["Team"], int(r["Season"])): (r["Off PPG"], r["Def PPG"])
        for r in ppg.to_dict("records")
    }


def build_efficiency_updates(raw: pd.DataFrame) -> dict[str, tuple[float, ...]]:
    """
    Pure function. Maps efficiency.compute_raw_efficiency()'s output to
    {Team: (epa_off, epa_def, success_off, success_def, nya_off, nya_def, proe)}, for
    writing into Advanced Efficiency Metrics Section 1.
    """
    return {r["Team"]: tuple(r[c] for c in EFFICIENCY_COLUMNS) for r in raw.to_dict("records")}


def _write_yoy_section(ws, updates: dict[tuple[str, int], tuple[float, float]]) -> int:
    """Writes Off/Def PPG into every (Team, Season) row of Section 1 until the first blank
    row. Returns the number of rows written."""
    written = 0
    row = 5
    while ws.cell(row=row, column=1).value is not None:
        team = ws.cell(row=row, column=1).value
        season = ws.cell(row=row, column=2).value
        key = (team, int(season))
        if key not in updates:
            raise KeyError(f"No pulled PPG data for {key} (row {row}) -- workbook/pull mismatch")
        off_ppg, def_ppg = updates[key]
        ws.cell(row=row, column=3, value=off_ppg)
        ws.cell(row=row, column=4, value=def_ppg)
        written += 1
        row += 1
    return written


def _write_efficiency_section(ws, updates: dict[str, tuple[float, ...]]) -> int:
    """Writes the 7 raw metric columns into every Team row of Section 1 until the first
    blank row. Returns the number of rows written."""
    written = 0
    row = 5
    while ws.cell(row=row, column=1).value is not None:
        team = ws.cell(row=row, column=1).value
        if team not in updates:
            raise KeyError(
                f"No pulled efficiency data for {team!r} (row {row}) -- workbook/pull mismatch"
            )
        for offset, value in enumerate(updates[team]):
            ws.cell(row=row, column=2 + offset, value=value)
        written += 1
        row += 1
    return written


def update_workbook(workbook_path: str, years: list[int] | None = None) -> dict[str, int]:
    """
    Pulls fresh nflverse data (schedules for PPG, play-by-play for efficiency metrics) and
    writes it into the workbook's two placeholder input tables, in place.
    Returns {"yoy_rows": n, "efficiency_rows": n}.
    """
    years = years or [2023, 2024, 2025]

    sched = fetch_schedules(years)
    ppg = transform_to_team_season(sched)

    pbp = fetch_pbp(years)
    raw_eff = compute_raw_efficiency(pbp)

    wb = openpyxl.load_workbook(workbook_path)
    yoy_written = _write_yoy_section(wb[YOY_SHEET], build_yoy_updates(ppg))
    eff_written = _write_efficiency_section(wb[EFFICIENCY_SHEET], build_efficiency_updates(raw_eff))
    wb.save(workbook_path)

    return {"yoy_rows": yoy_written, "efficiency_rows": eff_written}


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook_path", help="Path to the NFL Prediction Model .xlsx file")
    parser.add_argument(
        "--years", type=int, nargs="+", default=None,
        help="Seasons to pull, e.g. --years 2023 2024 2025 (default: 2023 2024 2025)",
    )
    args = parser.parse_args()
    result = update_workbook(args.workbook_path, args.years)
    print(
        f"Updated {result['yoy_rows']} rows in '{YOY_SHEET}' and "
        f"{result['efficiency_rows']} rows in '{EFFICIENCY_SHEET}'."
    )
    print("Open the workbook in Excel (or re-open it) to recalculate formulas.")


if __name__ == "__main__":
    main()
